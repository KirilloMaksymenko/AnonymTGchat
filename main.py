#project - AnonymTGchat , Author - Maksymenko Kyrylo

import random
import telebot
import openpyxl

class ExcelManager:

    def __init__(self,path):
        self.path = path
        self.wb = openpyxl.load_workbook(path)
        self.sheet = self.wb.active

    def update_settings(self,id,key,value):
        print(f"[INFO] >> Update settings {id}")
        self.sheet.cell(row= self.find_id(id) , column = self.read_settings(key)).value = value
        self.save_changes()

    def find_id(self,id):
        print(f"[INFO] >> Find id {id}")
        try:
            for i in range(1, self.sheet.max_row + 1):
                if str(self.sheet.cell(row = i, column = 1).value) == str(id):
                    return i
        except:
            print(f"[ERROR] << {Exception}")
            return None
        
    def read_settings(self,key):
        print(f"[INFO] >> Read settings {key}")
        for i in range(1, self.sheet.max_column + 1):
            if key in str(self.sheet.cell(row = 1, column = i).value):
                return i
    def create_new_persone(self,id):
        base_date = [str(id),"out","en","False","False","0","0","0"]
        pos_y = self.sheet.max_row + 1
        for i in range(1, self.sheet.max_column + 1):
            self.sheet.cell(row= pos_y , column = i).value = base_date[i - 1]
        self.save_changes()
        print(f"[INFO] >> Chat with id {id} add to database") 
    
    def save_changes(self):
        self.wb.save(self.path)


class TalkingManager(ExcelManager):
        
    def id_opponent(self,id):
        print(f"[INFO] >> Id opponent {self.sheet.cell(row = self.find_id(id), column = self.read_settings('opponent')).value}")
        return self.sheet.cell(row = self.find_id(id), column = self.read_settings('opponent')).value

    def search_opponent(self,id):
        list_ready = []
        for i in range(1, self.sheet.max_row + 1):
            if self.sheet.cell(row = i, column = self.read_settings("state")).value == "stay":
                list_ready.append(self.sheet.cell(row = i, column = self.read_settings("id")).value)
        if list_ready == []:
            return None
        opponent_id = random.choice(list_ready)
        self.sheet.cell(row= self.find_id(id) , column = self.read_settings("opponent")).value = str(opponent_id)
        self.sheet.cell(row= self.find_id(opponent_id) , column = self.read_settings("opponent")).value = str(id)
        self.talk_state(id)
        self.talk_state(opponent_id)
        self.save_changes()
        return opponent_id
    
    def check_state(self,id):
        return self.sheet.cell(row = self.find_id(id), column = self.read_settings("state")).value


    def talk_state(self,id):
        self.sheet.cell(row= self.find_id(id) , column = self.read_settings("state")).value = "talk"
        self.save_changes()

    def stay_state(self,id):
        self.sheet.cell(row= self.find_id(id) , column = self.read_settings("state")).value = "stay"
        self.save_changes()

    def out_state(self,id):
        self.sheet.cell(row= self.find_id(id) , column = self.read_settings("state")).value = "out"
        self.save_changes()

    def leave_chat(self,id):
        opponent_id = self.id_opponent(id)
        self.sheet.cell(row= self.find_id(id) , column = self.read_settings("opponent")).value = "0"
        self.sheet.cell(row= self.find_id(opponent_id) , column = self.read_settings("opponent")).value = "0"
        self.out_state(id)
        self.out_state(opponent_id)
        self.save_changes()
        return opponent_id


client = telebot.TeleBot("5683450613:AAGmLM6_jWWeuKISPBHBFVIBgcLbhc8UiV4")
talk = TalkingManager("source\dataset\data.xlsx")
#_____________CREATE_CELL_IN_DATABASE_________________#
@client.message_handler(commands=['start'])
def handle_start(message):
    if talk.find_id(message.chat.id) == None:
        talk.create_new_persone(message.chat.id)  
        

#_____________SETTINGS_________________#
@client.message_handler(commands=['set_settings'])
def settings_view(message):
    txt_info = f"1.Image allow - {talk.read_settings('img_allow')}\n2.Voice allow - {talk.read_settings('voice_allow')}"
    client.send_message(message.chat.id, txt_info)
    txt_help = "/img (True/False) Your opponent can send you img\n/voice (True/False) Your opponent can send you voice"
    client.send_message(message.chat.id, txt_help)

@client.message_handler(commands=['voice'])
def set_voice(message):
    param = str(message.text).replace("/voice ","")
    if param.title() in ["True","False"]:
        talk.update_settings(message.chat.id,"voice_allow",param)

@client.message_handler(commands=['image'])
def set_image(message):
    param = str(message.text).replace("/image ","")
    if param.title() in ["True","False"]:
        talk.update_settings(message.chat.id,"img_allow",param)
    
#__________Talking___________#

@client.message_handler(commands=['turn'])
def stay_in_turn(message):
    talk.stay_state(message.chat.id)
    client.send_message(message.chat.id,"You enter to turn")

@client.message_handler(commands=['search'])
def searche_opponent(message):
    opponent = talk.search_opponent(message.chat.id)
    if opponent == None:
        client.send_message(message.chat.id,"Try later")
        return
    print(opponent)
    client.send_message(message.chat.id,"You connect to talk\nIf you want to leave entre /leave")
    client.send_message(opponent,"You connect to talk\nIf you want to leave entre /leave")

@client.message_handler(commands=['leave'])
def leave_chat(message):
    opponent = talk.leave_chat(message.chat.id)
    client.send_message(message.chat.id,"You leave chat")
    client.send_message(opponent,"Your opponent leave chat")
    
@client.message_handler(func=lambda message: True)
def talking(message):
    if "talk" in talk.check_state(message.chat.id):
        opponent = talk.id_opponent(message.chat.id)
        client.send_message(opponent,f"Anonym >> {message.text}")





client.infinity_polling()